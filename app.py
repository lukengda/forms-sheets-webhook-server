import glob
import logging
import os
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from filelock import SoftFileLock
from flask import Flask, jsonify, request
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

app = Flask(__name__)

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=getattr(logging, LOG_LEVEL, logging.INFO))

FORM_RESPONSE_SHEET = "Form Responses"
FORM_RESPONSE_SHEET_AUTO_HEIGHT = "Form Responses (auto-height)"

# Output folder for the Excel workbooks. Defaults to the container mount
# point (/data); override with EXCEL_FOLDER to run locally, e.g. ./data.
EXCEL_FOLDER = os.getenv("EXCEL_FOLDER", "/data")
if not os.path.exists(EXCEL_FOLDER):
    raise Exception(
        f"Excel output folder {EXCEL_FOLDER!r} does not exist. "
        "Mount a volume there or set EXCEL_FOLDER to an existing directory."
    )

for f in glob.glob(EXCEL_FOLDER + "/*.lock"):
    app.logger.info(f"Deleting stale lock file on startup: {f}")
    os.remove(f)


def sanitize_filename(name: str):
    """Sanitize the file name to make it safe for saving."""
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def parse_form_data(form_data: dict):
    """Parse the form data to extract field names and values."""
    parsed_data = {}

    if "Submission Date" not in form_data:
        parsed_data["Submission Date"] = (
            datetime.now()
            .astimezone(ZoneInfo("Europe/Berlin"))
            .strftime("%Y-%m-%d %H:%M:%S")
        )

    for field_id, field_content in form_data.get("data", {}).items():
        field_name = field_content.get("name", f"Field_{field_id}")  # Use fallback name
        field_value = field_content.get("value", "")
        if isinstance(field_value, str):
            try:
                # Convert to universal excel-compatible datetime format if format matches
                parsed_date = datetime.strptime(
                    field_value, "%d/%m/%Y %H:%M"
                ).astimezone(ZoneInfo("Europe/Berlin"))
                field_value = parsed_date.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass  # Ignore non-date strings
        parsed_data[field_name] = field_value
    return parsed_data


def get_excel_path(form_title: str):
    """Get sanitized Excel file path for the form title."""
    safe_title = sanitize_filename(form_title)
    return os.path.join(EXCEL_FOLDER, f"{safe_title}.xlsx")


def get_response_sheet(workbook: Workbook) -> Worksheet:
    if FORM_RESPONSE_SHEET_AUTO_HEIGHT in workbook.sheetnames:
        return workbook[FORM_RESPONSE_SHEET_AUTO_HEIGHT]
    if FORM_RESPONSE_SHEET not in workbook.sheetnames:
        app.logger.debug(f"Creating sheet: {FORM_RESPONSE_SHEET}")
        workbook.create_sheet(FORM_RESPONSE_SHEET)

    return workbook[FORM_RESPONSE_SHEET]


def ensure_excel_file(excel_path: str) -> Workbook:
    """Ensure an Excel file exists for the form with the necessary sheet."""
    if os.path.exists(excel_path):
        app.logger.debug(f"Loading existing workbook: {excel_path}")
        return load_workbook(excel_path)

    app.logger.info(f"Creating new workbook: {excel_path}")
    return Workbook()


def ensure_headers(sheet: Worksheet, data: dict):
    """Ensure the headers exist in the response sheet."""
    headers = data.keys()
    current_headers = {str(cell.value) for cell in sheet[1] if cell.value}
    new_headers = [header for header in headers if header not in current_headers]

    # Add new headers at the end of the current headers
    for col, header in enumerate(new_headers, start=len(current_headers) + 1):
        sheet.cell(row=1, column=col, value=header)


def append_data(workbook: Workbook, data: dict):
    """Append data to the first free row in the response sheet."""
    response_sheet: Worksheet = get_response_sheet(workbook)
    headers = {
        cell.value: col
        for col, cell in enumerate(response_sheet[1], start=1)
        if cell.value
    }

    new_row_idx = response_sheet.max_row + 1
    for header, col in headers.items():
        response_sheet.cell(row=new_row_idx, column=col, value=data.get(header, ""))

    # Adjust row height dynamically
    for sheet in workbook.worksheets:
        app.logger.debug(f"{sheet.title=}")
        if sheet.title.endswith("(auto-height)"):
            app.logger.debug(f"{sheet.title.endswith('(auto-height)')=}")
            for row in sheet.iter_rows(min_row=new_row_idx, max_row=new_row_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping
                app.logger.debug("Enabling auto height for row %r", row[0].row)
                sheet.row_dimensions[row[0].row].height = None


@app.route("/webhook", methods=["POST"])
def handle_webhook():
    """Handle incoming webhooks and save data to an Excel file."""
    try:
        # Extract JSON payload
        if not request.json:
            return jsonify({"error": "No data provided"}), 400
        form_title = request.json["form_title"]
        form_data = parse_form_data(request.json)
        if not form_data:
            return jsonify({"error": "No data provided"}), 400

        app.logger.debug(f"{form_data=}")

        excel_path = get_excel_path(form_title)
        with SoftFileLock(excel_path + ".lock"):
            workbook = ensure_excel_file(excel_path)

            ensure_headers(get_response_sheet(workbook), form_data)
            append_data(workbook, form_data)
            workbook.save(excel_path)

            app.logger.info(f"Data saved successfully in {excel_path}")
            return jsonify({"message": "Data saved successfully"}), 200
    except Exception as e:
        app.logger.exception("Unhandled error: %r", e)
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
